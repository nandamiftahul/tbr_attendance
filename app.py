# app.py

import os
import math
from datetime import datetime, date, time, timedelta
from zoneinfo import ZoneInfo

from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from dotenv import load_dotenv
from io import StringIO
import csv, io
import openpyxl
from openpyxl.utils import get_column_letter

from sqlalchemy import or_

from models import db, User, Employee, Attendance, LeaveRequest, Holiday, Shift, Office, Announcement
from flask import jsonify, Response
from flask_login import login_required
from werkzeug.security import generate_password_hash

GOOGLE_ENABLED = False
try:
    from authlib.integrations.flask_client import OAuth
except Exception:
    OAuth = None

load_dotenv()
TZ = os.getenv("TIMEZONE", "Asia/Jakarta")

app = Flask(__name__, instance_relative_config=True)
app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "dev-secret")
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv("SQLALCHEMY_DATABASE_URI", "sqlite:///app.sqlite")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

os.makedirs(app.instance_path, exist_ok=True)

db.init_app(app)

login_manager = LoginManager(app)
login_manager.login_view = "login"

@login_manager.unauthorized_handler
def unauthorized_callback():
    # For API calls, return JSON instead of HTML redirect
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": "Unauthorized"}), 401
    return redirect(url_for("login", next=request.path))

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def get_client_ip() -> str:
    return (request.headers.get("X-Forwarded-For", request.remote_addr or "") or "")[:63]

def get_user_agent() -> str:
    return (request.headers.get("User-Agent") or "")[:255]

def haversine_m(lat1, lon1, lat2, lon2) -> float:
    R = 6371000.0
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))

def geofence_ok(lat, lon) -> bool:
    office = Office.query.filter_by(is_active=True).first()
    if not office:
        return True

    if lat is None or lon is None:
        return False

    return haversine_m(
        office.lat, office.lon,
        float(lat), float(lon)
    ) <= office.radius_m

def active_shift_for(emp: Employee) -> Shift:
    if emp and emp.shift:
        return emp.shift
    s = Shift.query.filter_by(name="Office").first()
    if s:
        return s
    return Shift(name="Default", start_time=time(9, 0), end_time=time(17, 0), grace_in_min=15, grace_out_min=0)

def is_holiday(d: date) -> bool:
    return Holiday.query.filter_by(date=d).first() is not None

def approved_leave_for(emp_id: int, d: date):
    return LeaveRequest.query.filter(
        LeaveRequest.employee_id == emp_id,
        LeaveRequest.status == "approved",
        LeaveRequest.start_date <= d,
        LeaveRequest.end_date >= d,
    ).first()

def to_local_iso(dt):
    if not dt:
        return ""
    tz = ZoneInfo(TZ)
    # if dt naive, anggap UTC (paling aman di server UTC) lalu convert
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=ZoneInfo("UTC"))
    return dt.astimezone(tz).isoformat()


# ✅ helper create/link user
def create_user_for_employee(emp: Employee, login_email: str, password: str, role: str = "staff"):
    login_email = (login_email or "").strip().lower()
    if not login_email:
        raise ValueError("Login email/username is required.")
    if not password:
        raise ValueError("Password is required.")
    if User.query.filter_by(email=login_email).first():
        raise ValueError("Login email already exists.")

    # Normalize role
    role = (role or "staff").strip().lower()
    if role not in ("admin", "staff", "manager", "general_manager", "hrd"):
        role = "staff"

    u = User(name=emp.name, email=login_email, role=role, is_active=True)
    u.set_password(password)
    db.session.add(u)
    db.session.flush()  # get u.id
    emp.user_id = u.id


@app.cli.command("initdb")
def initdb():
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(email="admin@example.com").first():
            admin = User(name="Administrator", email="admin@example.com", role="admin")
            admin.set_password("admin123")
            db.session.add(admin)
        if Shift.query.count() == 0:
            db.session.add(Shift(name="Office", start_time=time(9, 0), end_time=time(17, 0), grace_in_min=15, grace_out_min=0))
        if Office.query.count() == 0:
            db.session.add(Office(
                name="Test Office",
                lat=-6.3614579,
                lon=106.8180971,
                radius_m=300,
                is_active=True
            ))
        db.session.commit()
        print("DB initialized. Default admin: admin@example.com / admin123")

@app.cli.command("mark-absent")
def mark_absent():
    with app.app_context():
        today = datetime.now(ZoneInfo(TZ)).date()
        if is_holiday(today):
            print("Holiday today. Skipping absent marking.")
            return
        active = Employee.query.filter_by(is_active=True).all()
        created = 0
        for e in active:
            if approved_leave_for(e.id, today):
                continue
            exists = Attendance.query.filter_by(employee_id=e.id, work_date=today).first()
            if not exists:
                db.session.add(Attendance(employee_id=e.id, work_date=today, status="absent", note="Auto-marked"))
                created += 1
        db.session.commit()
        print(f"Absent records created: {created}")

# ---- Google SSO part unchanged ----
if OAuth and os.getenv("GOOGLE_OAUTH_CLIENT_ID"):
    oauth = OAuth(app)
    oauth.register(
        name="google",
        client_id=os.getenv("GOOGLE_OAUTH_CLIENT_ID"),
        client_secret=os.getenv("GOOGLE_OAUTH_CLIENT_SECRET"),
        server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
        client_kwargs={"scope": "openid email profile"},
    )
    GOOGLE_ENABLED = True

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = (request.form.get("email") or "").strip().lower()
        password = request.form.get("password") or ""
        user = User.query.filter_by(email=email, is_active=True).first()
        if user and user.check_password(password):
            login_user(user)
            flash("Welcome back!", "success")
            return redirect(url_for("dashboard"))
        flash("Invalid credentials", "danger")
    return render_template("login.html", google_enabled=GOOGLE_ENABLED)

@app.get("/login/google")
def login_google():
    if not GOOGLE_ENABLED:
        flash("Google SSO is disabled", "warning")
        return redirect(url_for("login"))
    redirect_uri = url_for("auth_callback_google", _external=True)
    return oauth.google.authorize_redirect(redirect_uri)

@app.get("/auth/callback/google")
def auth_callback_google():
    if not GOOGLE_ENABLED:
        return redirect(url_for("login"))
    token = oauth.google.authorize_access_token()
    userinfo = token.get("userinfo") or oauth.google.parse_id_token(token)
    email = (userinfo.get("email") or "").strip().lower()
    name = userinfo.get("name") or (email.split("@")[0] if email else "User")
    user = User.query.filter_by(email=email).first()
    if not user:
        user = User(name=name, email=email, role="staff", is_active=True)
        user.set_password(os.urandom(8).hex())
        db.session.add(user)
        db.session.commit()
    login_user(user)
    flash("Logged in with Google", "success")
    return redirect(url_for("dashboard"))

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Logged out", "info")
    return redirect(url_for("login"))


def _employee_for_user(user: User):
    # employee linked via employees.user_id
    return Employee.query.filter_by(user_id=user.id).first()

@app.post("/api/profile/change_password")
@login_required
def api_change_password():
    data = request.get_json(silent=True) or {}
    old_pw = (data.get("old_password") or "").strip()
    new_pw = (data.get("new_password") or "").strip()

    if not old_pw or not new_pw:
        return jsonify({"ok": False, "error": "Old and new password required"}), 400
    if len(new_pw) < 6:
        return jsonify({"ok": False, "error": "Password must be at least 6 characters"}), 400

    # check old password
    if not current_user.check_password(old_pw):
        return jsonify({"ok": False, "error": "Old password is wrong"}), 400

    current_user.set_password(new_pw)
    db.session.commit()
    return jsonify({"ok": True, "message": "Password updated"})

@app.post("/api/login")
def api_login():
    data = request.get_json(force=True, silent=True) or {}
    email = (data.get("email") or "").strip().lower()
    password = data.get("password") or ""
    user = User.query.filter_by(email=email, is_active=True).first()
    if not user or not user.check_password(password):
        return jsonify({"ok": False, "error": "Invalid credentials"}), 401

    # create a very simple tokenless session by logging-in (cookie session)
    login_user(user)
    return jsonify({"ok": True, "role": user.role, "name": user.name, "email": user.email})

@app.post("/api/logout")
@login_required
def api_logout():
    logout_user()
    return jsonify({"ok": True})

@app.get("/api/me")
@login_required
def api_me():
    emp = _employee_for_user(current_user)
    return jsonify({
        "ok": True,
        "user": {"name": current_user.name, "email": current_user.email, "role": current_user.role},
        "employee": ({"id": emp.id, "code": emp.code, "name": emp.name, "dept": emp.dept} if emp else None)
    })

@app.post("/api/attendance/check")
@login_required
def api_attendance_check():
    emp = _employee_for_user(current_user)
    if not emp:
        return jsonify({"ok": False, "error": "No employee linked to this account"}), 400

    data = request.get_json(force=True, silent=True) or {}
    action = data.get("action")  # check_in/check_out
    lat = data.get("lat")
    lon = data.get("lon")

    lat = float(lat) if lat is not None else None
    lon = float(lon) if lon is not None else None

    now = datetime.now(ZoneInfo(TZ))
    today = now.date()

    # optional geofence on check-in
    if action == "check_in" and not geofence_ok(lat, lon):
        return jsonify({"ok": False, "error": "Outside geofence"}), 403

    record = Attendance.query.filter_by(employee_id=emp.id, work_date=today).first()
    if not record:
        record = Attendance(employee_id=emp.id, work_date=today)
        db.session.add(record)

    if action == "check_in":
        if record.check_in:
            return jsonify({"ok": False, "error": "Already checked in"}), 409
        record.check_in = now
        record.check_in_ip = get_client_ip()
        record.check_in_ua = get_user_agent()
        record.check_in_lat = lat
        record.check_in_lon = lon
        record.status = "present"
        db.session.commit()
        return jsonify({"ok": True, "message": "Checked in", "time": now.isoformat()})

    if action == "check_out":
        if not record.check_in:
            return jsonify({"ok": False, "error": "Cannot check out before check in"}), 409
        if record.check_out:
            return jsonify({"ok": False, "error": "Already checked out"}), 409
        record.check_out = now
        record.check_out_ip = get_client_ip()
        record.check_out_ua = get_user_agent()
        record.check_out_lat = lat
        record.check_out_lon = lon
        db.session.commit()
        return jsonify({"ok": True, "message": "Checked out", "time": now.isoformat()})

    return jsonify({"ok": False, "error": "Unknown action"}), 400

@app.post("/api/leave/request")
@login_required
def api_leave_request():
    emp = _employee_for_user(current_user)
    if not emp:
        return jsonify({"ok": False, "error": "No employee linked"}), 400

    data = request.get_json(silent=True) or {}

    req_type = (data.get("type") or "").strip().lower()
    reason = (data.get("reason") or "").strip()

    try:
        start_date = datetime.strptime(data.get("start_date"), "%Y-%m-%d").date()
        end_date = datetime.strptime(data.get("end_date"), "%Y-%m-%d").date()
    except Exception:
        return jsonify({"ok": False, "error": "Invalid date format"}), 400

    if req_type not in ("leave", "sick", "wfh", "on_site"):
        return jsonify({"ok": False, "error": "Invalid type"}), 400

    if end_date < start_date:
        return jsonify({"ok": False, "error": "End date must be >= start date"}), 400

    # Determine initial status based on requester role
    req_role = (current_user.role or "staff").lower()
    if req_role in ("admin", "hrd"):
        init_status = "approved"
    elif req_role in ("manager", "general_manager"):
        # manager's own request goes directly to HRD
        init_status = "pending_hrd"
    else:
        init_status = "pending_manager"

    r = LeaveRequest(
        employee_id=emp.id,
        type=req_type,
        start_date=start_date,
        end_date=end_date,
        reason=reason,
        status=init_status,
    )

    db.session.add(r)
    db.session.commit()

    return jsonify({
        "ok": True,
        "id": r.id,
        "status": r.status
    })


# --- API: list my leave requests ---
@app.get("/api/leave/my")
@login_required
def api_leave_my():
    emp = _employee_for_user(current_user)
    if not emp:
        return jsonify({"ok": False, "error": "No employee linked"}), 400

    rows = LeaveRequest.query.filter_by(employee_id=emp.id).order_by(LeaveRequest.id.desc()).all()
    return jsonify({"ok": True, "rows": [{
        "id": r.id,
        "type": r.type,
        "start_date": str(r.start_date),
        "end_date": str(r.end_date),
        "status": r.status,
        "reason": r.reason or "",
        "manager_approved_by": getattr(r, "manager_approved_by", None),
        "hrd_approved_by": getattr(r, "hrd_approved_by", None),
        "approved_by": getattr(r, "approved_by", None),
    } for r in rows]})


# --- API: list my attendance records (FULL REPLACE) ---
@app.get("/api/attendance/my")
@login_required
def api_attendance_my():
    emp = _employee_for_user(current_user)
    if not emp:
        return jsonify({"ok": False, "error": "No employee linked"}), 400

    # ---- helpers ----
    def parse_ymd(s: str):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    def to_local_iso(dt):
        """
        Return ISO8601 with timezone offset (+07:00 for Asia/Jakarta).
        If DB stored naive datetime, assume it's UTC then convert to local.
        """
        if not dt:
            return ""
        tz_local = ZoneInfo(TZ)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        return dt.astimezone(tz_local).isoformat()

    # ---- query params ----
    start_s = (request.args.get("start") or "").strip()  # YYYY-MM-DD
    end_s = (request.args.get("end") or "").strip()      # YYYY-MM-DD

    start_d = parse_ymd(start_s) if start_s else None
    end_d = parse_ymd(end_s) if end_s else None

    if start_s and not start_d:
        return jsonify({"ok": False, "error": "Invalid start date (use YYYY-MM-DD)"}), 400
    if end_s and not end_d:
        return jsonify({"ok": False, "error": "Invalid end date (use YYYY-MM-DD)"}), 400

    q = Attendance.query.filter_by(employee_id=emp.id)

    if start_d:
        q = q.filter(Attendance.work_date >= start_d)
    if end_d:
        q = q.filter(Attendance.work_date <= end_d)

    rows = q.order_by(Attendance.work_date.desc()).limit(500).all()

    return jsonify({
        "ok": True,
        "rows": [{
            "date": r.work_date.isoformat(),
            # ✅ return ISO with timezone offset for mobile correctness
            "check_in": to_local_iso(r.check_in),
            "check_out": to_local_iso(r.check_out),
            "hours": float(r.duration_hours or 0.0),
            "status": r.status,
            "note": r.note or "",
        } for r in rows]
    })


# --- API: holidays (for Records calendar) ---
@app.get("/api/holidays")
@login_required
def api_holidays():
    """
    Return holiday list in a date range.
    Query: /api/holidays?start=YYYY-MM-DD&end=YYYY-MM-DD
    """
    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()

    def parse_ymd(s: str):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    start_d = parse_ymd(start_s) if start_s else None
    end_d = parse_ymd(end_s) if end_s else None
    if not start_d or not end_d:
        return jsonify({"ok": False, "error": "start/end required (YYYY-MM-DD)"}), 400
    if end_d < start_d:
        return jsonify({"ok": False, "error": "end must be >= start"}), 400

    rows = (Holiday.query
            .filter(Holiday.date >= start_d, Holiday.date <= end_d)
            .order_by(Holiday.date.asc())
            .all())

    return jsonify({
        "ok": True,
        "rows": [{"date": h.date.isoformat(), "name": h.name} for h in rows]
    })


@app.get("/api/announcements/active")
@login_required
def api_announcements_active():
    now = datetime.utcnow()
    q = Announcement.query.filter_by(is_active=True)

    # jika start/end dipakai: aktif jika (start null or start<=now) and (end null or end>=now)
    q = q.filter(
        (Announcement.start_at.is_(None) | (Announcement.start_at <= now)),
        (Announcement.end_at.is_(None) | (Announcement.end_at >= now)),
    )

    rows = q.order_by(Announcement.created_at.desc()).limit(10).all()

    def to_row(a: Announcement):
        return {
            "id": a.id,
            "title": a.title,
            "body": a.body,
            "level": (a.level or "info"),
            "created_at": a.created_at.isoformat() if a.created_at else "",
        }

    return jsonify({"ok": True, "rows": [to_row(a) for a in rows]})


# ---------------------------
# API: approvals for mobile
# ---------------------------

def _leave_to_row(r: LeaveRequest):
    # r.employee should exist via relationship (LeaveRequest -> Employee)
    emp = getattr(r, "employee", None)
    return {
        "id": r.id,
        "type": r.type,
        "start_date": str(r.start_date),
        "end_date": str(r.end_date),
        "status": r.status,
        "reason": r.reason or "",
        "employee_id": r.employee_id,
        "employee_name": (emp.name if emp else ""),
        "dept": (emp.dept if emp else ""),
        "manager_approved_by": getattr(r, "manager_approved_by", None),
        "hrd_approved_by": getattr(r, "hrd_approved_by", None),
    }


def _mark_attendance_for_leave(r: LeaveRequest):
    d = r.start_date
    while d <= r.end_date:
        a = Attendance.query.filter_by(employee_id=r.employee_id, work_date=d).first()
        if not a:
            a = Attendance(employee_id=r.employee_id, work_date=d)
            db.session.add(a)

        if r.type == "wfh":
            a.status = "wfh"
        elif r.type == "on_site":
            a.status = "on_site"
        else:
            a.status = r.type  # leave / sick

        a.note = (a.note or "") + f" Marked by approval #{r.id}."
        d += timedelta(days=1)


def _my_dept():
    me_emp = _employee_for_user(current_user)
    return (me_emp.dept if me_emp else None)


@app.get("/api/leave/approvals")
@login_required
def api_leave_approvals():
    role = (current_user.role or "staff").lower()

    # Staff: tidak ada approvals (biar tab kosong)
    if role == "staff":
        return jsonify({"ok": True, "rows": []})

    # Manager: pending_manager untuk dept dia saja
    if role == "manager":
        dept = _my_dept()
        if not dept:
            return jsonify({"ok": True, "rows": []})
        rows = (LeaveRequest.query
                .join(Employee)
                .filter(Employee.dept == dept)
                .filter(LeaveRequest.status == "pending_manager")
                .order_by(LeaveRequest.id.desc())
                .all())
        return jsonify({"ok": True, "rows": [_leave_to_row(r) for r in rows]})

    # General Manager: pending_manager semua dept
    if role == "general_manager":
        rows = (LeaveRequest.query
                .filter(LeaveRequest.status == "pending_manager")
                .order_by(LeaveRequest.id.desc())
                .all())
        return jsonify({"ok": True, "rows": [_leave_to_row(r) for r in rows]})

    # HRD: pending_hrd saja
    if role == "hrd":
        rows = (LeaveRequest.query
                .filter(LeaveRequest.status == "pending_hrd")
                .order_by(LeaveRequest.id.desc())
                .all())
        return jsonify({"ok": True, "rows": [_leave_to_row(r) for r in rows]})

    # Admin: lihat semua yang belum final (atau semua kalau mau)
    if role == "admin":
        rows = (LeaveRequest.query
                .filter(LeaveRequest.status.in_(["pending_manager", "pending_hrd"]))
                .order_by(LeaveRequest.id.desc())
                .all())
        return jsonify({"ok": True, "rows": [_leave_to_row(r) for r in rows]})

    return jsonify({"ok": False, "error": "Not allowed"}), 403


@app.post("/api/leave/<int:rid>/approve")
@login_required
def api_leave_approve(rid):
    role = (current_user.role or "staff").lower()
    r = LeaveRequest.query.get_or_404(rid)

    # Manager stage
    if role in ("manager", "general_manager"):
        if r.status != "pending_manager":
            return jsonify({"ok": False, "error": "Not pending manager approval"}), 409

        if role == "manager":
            dept = _my_dept()
            if not dept or not r.employee or r.employee.dept != dept:
                return jsonify({"ok": False, "error": "Not allowed (different department)"}), 403

        r.status = "pending_hrd"
        r.manager_approved_by = current_user.id
        r.manager_approved_at = datetime.utcnow()
        db.session.commit()
        return jsonify({"ok": True, "status": r.status, "row": _leave_to_row(r)})

    # HRD final stage (admin juga boleh final approve)
    if role in ("hrd", "admin"):
        if r.status not in ("pending_hrd", "pending_manager"):
            return jsonify({"ok": False, "error": "Not pending HRD approval"}), 409

        r.status = "approved"
        r.hrd_approved_by = current_user.id
        r.hrd_approved_at = datetime.utcnow()

        _mark_attendance_for_leave(r)
        db.session.commit()
        return jsonify({"ok": True, "status": r.status, "row": _leave_to_row(r)})

    return jsonify({"ok": False, "error": "Not allowed"}), 403


@app.post("/api/leave/<int:rid>/reject")
@login_required
def api_leave_reject(rid):
    role = (current_user.role or "staff").lower()
    r = LeaveRequest.query.get_or_404(rid)

    allowed = False

    # Admin: boleh reject kapan saja
    if role == "admin":
        allowed = True

    # Manager/GM: boleh reject hanya pending_manager
    elif role in ("manager", "general_manager") and r.status == "pending_manager":
        allowed = True
        if role == "manager":
            dept = _my_dept()
            if not dept or not r.employee or r.employee.dept != dept:
                allowed = False

    # HRD: boleh reject hanya pending_hrd
    elif role == "hrd" and r.status == "pending_hrd":
        allowed = True

    if not allowed:
        return jsonify({"ok": False, "error": "Not allowed"}), 403

    r.status = "rejected"
    r.rejected_by = current_user.id
    r.rejected_at = datetime.utcnow()
    db.session.commit()
    return jsonify({"ok": True, "status": r.status, "row": _leave_to_row(r)})


@app.route("/")
@login_required
def dashboard():
    today = datetime.now(ZoneInfo(TZ)).date()
    recs = Attendance.query.filter_by(work_date=today).all()
    present = sum(1 for r in recs if r.status in ("present", "late", "wfh", "on_site"))
    leave = sum(1 for r in recs if r.status == "leave")
    sick = sum(1 for r in recs if r.status == "sick")
    wfh = sum(1 for r in recs if r.status == "wfh")
    absent = sum(1 for r in recs if r.status == "absent")
    total = len(recs)
    return render_template("dashboard.html", today=today, stats=dict(total=total, present=present, leave=leave, sick=sick, wfh=wfh, absent=absent))

@app.route("/employees")
@login_required
def employees():
    if current_user.role != "admin":
        flash("Admins only", "warning")
        return redirect(url_for("dashboard"))
    q = request.args.get("q", "").strip()
    query = Employee.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Employee.name.ilike(like), Employee.email.ilike(like), Employee.code.ilike(like), Employee.dept.ilike(like)))
    rows = query.order_by(Employee.name.asc()).all()
    shifts = Shift.query.order_by(Shift.start_time.asc()).all()
    return render_template("employees.html", rows=rows, q=q, shifts=shifts)

@app.post("/employees/create")
@login_required
def employee_create():
    if current_user.role != "admin":
        flash("Admins only", "warning")
        return redirect(url_for("employees"))

    shift_id = request.form.get("shift_id")
    emp = Employee(
        code=request.form.get("code"),
        name=request.form.get("name"),
        email=request.form.get("email"),
        dept=request.form.get("dept"),
        role=(request.form.get("role") or "staff").strip().lower(),
        shift_id=(int(shift_id) if shift_id else None),
        is_active=True,
    )
    db.session.add(emp)
    db.session.flush()

    # ✅ optional create login
    if request.form.get("create_login") == "1":
        login_email = request.form.get("login_email")
        login_password = request.form.get("login_password")
        login_role = (request.form.get("role") or emp.role or "staff").strip().lower()
        try:
            create_user_for_employee(emp, login_email, login_password, role=login_role)
            flash(f"Login created: {login_email}", "info")
        except Exception as e:
            db.session.rollback()
            flash(f"Employee created but login failed: {e}", "danger")
            return redirect(url_for("employees"))

    db.session.commit()
    flash("Employee created", "success")
    return redirect(url_for("employees"))

@app.post("/employees/<int:emp_id>/update")
@login_required
def employee_update(emp_id):
    if current_user.role != "admin":
        flash("Admins only", "warning")
        return redirect(url_for("employees"))

    emp = Employee.query.get_or_404(emp_id)
    emp.code = request.form.get("code")
    emp.name = request.form.get("name")
    emp.email = request.form.get("email")
    emp.dept = request.form.get("dept")
    emp.role = (request.form.get("role") or emp.role or "staff").strip().lower()
    emp.is_active = bool(request.form.get("is_active"))

    shift_id = request.form.get("shift_id")
    emp.shift_id = (int(shift_id) if shift_id else None)

    # ✅ login update / create / reset
    login_email = (request.form.get("login_email") or "").strip().lower()
    reset_pw = request.form.get("reset_password") == "1"
    new_pw = request.form.get("new_password") or ""

    try:
        if emp.user_id:
            u = User.query.get(emp.user_id)
            if u:
                if login_email and login_email != u.email:
                    if User.query.filter(User.email == login_email, User.id != u.id).first():
                        raise ValueError("Login email already exists.")
                    u.email = login_email
                u.name = emp.name
                u.role = emp.role or u.role
                u.is_active = emp.is_active

                if reset_pw:
                    if not new_pw:
                        raise ValueError("New password cannot be empty.")
                    u.set_password(new_pw)
        else:
            # create login from modal if requested
            if request.form.get("create_login") == "1":
                if not login_email:
                    raise ValueError("Login email is required.")
                if not new_pw:
                    raise ValueError("Password is required.")
                create_user_for_employee(emp, login_email, new_pw, role=(emp.role or 'staff'))

        db.session.commit()
        flash("Employee updated", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Update failed: {e}", "danger")

    return redirect(url_for("employees"))

@app.post("/employees/<int:emp_id>/delete")
@login_required
def employee_delete(emp_id):
    if current_user.role != "admin":
        flash("Admins only", "warning")
        return redirect(url_for("employees"))

    emp = Employee.query.get_or_404(emp_id)
    try:
        # optional: delete linked user
        if emp.user_id:
            u = User.query.get(emp.user_id)
            if u:
                db.session.delete(u)
        db.session.delete(emp)
        db.session.commit()
        flash("Employee deleted", "info")
    except Exception as e:
        db.session.rollback()
        flash(f"Delete failed: {e}", "danger")
    return redirect(url_for("employees"))

@app.post("/shifts/create")
@login_required
def shift_create():
    if current_user.role != "admin":
        return redirect(url_for("employees"))
    name = request.form.get("name")
    st = datetime.strptime(request.form.get("start_time"), "%H:%M").time()
    et = datetime.strptime(request.form.get("end_time"), "%H:%M").time()
    grace_in = int(request.form.get("grace_in_min", 15))
    grace_out = int(request.form.get("grace_out_min", 0))
    s = Shift(name=name, start_time=st, end_time=et, grace_in_min=grace_in, grace_out_min=grace_out)
    db.session.add(s)
    db.session.commit()
    flash("Shift created", "success")
    return redirect(url_for("employees"))

@app.post("/holidays/create")
@login_required
def holiday_create():
    if current_user.role != "admin":
        return redirect(url_for("dashboard"))
    d = datetime.strptime(request.form.get("date"), "%Y-%m-%d").date()
    name = request.form.get("name")
    if not Holiday.query.filter_by(date=d).first():
        db.session.add(Holiday(date=d, name=name))
        db.session.commit()
        flash("Holiday added", "success")
    else:
        flash("Holiday already exists", "warning")
    return redirect(url_for("dashboard"))

@app.get("/attendance/check")
@login_required
def attendance_check_get():
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.name.asc()).all()
    return render_template("attendance_check.html", employees=employees)

@app.post("/attendance/check")
@login_required
def attendance_check_post():
    emp_id = int(request.form.get("employee_id"))
    action = request.form.get("action")
    lat = request.form.get("lat")
    lon = request.form.get("lon")
    lat = float(lat) if lat else None
    lon = float(lon) if lon else None
    now = datetime.now(ZoneInfo(TZ))
    today = now.date()
    emp = Employee.query.get_or_404(emp_id)

    if action == "check_in" and not geofence_ok(lat, lon):
        flash("Check-in rejected: outside geofence.", "danger")
        return redirect(url_for("attendance_check_get"))

    record = Attendance.query.filter_by(employee_id=emp_id, work_date=today).first()
    if not record:
        record = Attendance(employee_id=emp_id, work_date=today)
        db.session.add(record)

    if action == "check_in":
        if record.check_in:
            flash("Already checked in", "warning")
        else:
            record.check_in = now
            record.check_in_ip = get_client_ip()
            record.check_in_ua = get_user_agent()
            record.check_in_lat = lat
            record.check_in_lon = lon

            sh = active_shift_for(emp)
            start_dt = datetime.combine(today, sh.start_time, tzinfo=ZoneInfo(TZ))
            late_threshold = start_dt + timedelta(minutes=sh.grace_in_min)
            if now > late_threshold:
                record.status = "late"
                record.note = (record.note or "") + f" Late check-in; shift {sh.name}."
            else:
                record.status = "present"
            flash("Checked in", "success")

    elif action == "check_out":
        if not record.check_in:
            flash("Cannot check out before check in", "danger")
        elif record.check_out:
            flash("Already checked out", "warning")
        else:
            record.check_out = now
            record.check_out_ip = get_client_ip()
            record.check_out_ua = get_user_agent()
            record.check_out_lat = lat
            record.check_out_lon = lon

            sh = active_shift_for(emp)
            end_dt = datetime.combine(today, sh.end_time, tzinfo=ZoneInfo(TZ))
            early_threshold = end_dt - timedelta(minutes=sh.grace_out_min)
            if now < early_threshold:
                record.note = (record.note or "") + f" Early checkout; shift {sh.name}."
            flash("Checked out — good job!", "success")
    else:
        flash("Unknown action", "danger")

    db.session.commit()
    return redirect(url_for("attendance_check_get"))

# --- WEB: Attendance list (FULL REPLACE) ---
@app.get('/attendance/list')
@login_required
def attendance_list():
    start_s = (request.args.get('start') or '').strip()
    end_s = (request.args.get('end') or '').strip()
    emp_q = (request.args.get('employee') or '').strip()

    def parse_ymd(s: str):
        try:
            return datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    tz_local = ZoneInfo(TZ)

    def to_local_dt(dt):
        """
        If dt is naive, assume UTC then convert to local.
        If dt has tzinfo, just convert to local.
        """
        if not dt:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        return dt.astimezone(tz_local)

    start_d = parse_ymd(start_s) if start_s else None
    end_d = parse_ymd(end_s) if end_s else None

    query = Attendance.query.join(Employee)

    if start_d:
        query = query.filter(Attendance.work_date >= start_d)
    if end_d:
        query = query.filter(Attendance.work_date <= end_d)
    if emp_q:
        query = query.filter(Employee.name.ilike(f"%{emp_q}%"))

    rows = query.order_by(Attendance.work_date.desc(), Employee.name.asc()).all()

    # Attach pre-formatted local strings so template stays simple
    for r in rows:
        ci = to_local_dt(r.check_in)
        co = to_local_dt(r.check_out)
        r.check_in_local_str = ci.strftime("%Y-%m-%d %H:%M:%S") if ci else ""
        r.check_out_local_str = co.strftime("%Y-%m-%d %H:%M:%S") if co else ""

    return render_template(
        'attendance_list.html',
        rows=rows,
        start=start_s,
        end=end_s,
        emp=emp_q
    )

@app.get('/attendance/export.csv')
@login_required
def attendance_export_csv():
    start = request.args.get('start')
    end = request.args.get('end')
    emp = request.args.get('employee')

    query = Attendance.query.join(Employee)
    if start:
        query = query.filter(Attendance.work_date >= start)
    if end:
        query = query.filter(Attendance.work_date <= end)
    if emp:
        query = query.filter(Employee.name.ilike(f"%{emp}%"))

    rows = query.order_by(Attendance.work_date.asc(), Employee.name.asc()).all()

    output = io.StringIO()
    cw = csv.writer(output)
    cw.writerow(['Date', 'Employee Code', 'Employee Name', 'Dept',
                 'Check-In', 'Check-Out', 'Hours', 'Status', 'Note',
                 'CheckIn IP', 'CheckIn UA', 'CheckOut IP', 'CheckOut UA'])

    for r in rows:
        cw.writerow([
            r.work_date.isoformat(),
            r.employee.code,
            r.employee.name,
            r.employee.dept or '',
            to_local_iso(r.check_in),
            to_local_iso(r.check_out),
            r.duration_hours,
            r.status,
            r.note or '',
            r.check_in_ip or '',
            r.check_in_ua or '',
            r.check_out_ip or '',
            r.check_out_ua or '',
        ])

    csv_data = output.getvalue()
    return Response(
        csv_data,
        mimetype='text/csv',
        headers={"Content-Disposition": "attachment; filename=attendance_export.csv"}
    )

@app.get('/attendance/export.xlsx')
@login_required
def attendance_export_xlsx():
    start = request.args.get('start')
    end = request.args.get('end')
    emp = request.args.get('employee')

    query = Attendance.query.join(Employee)
    if start:
        query = query.filter(Attendance.work_date >= start)
    if end:
        query = query.filter(Attendance.work_date <= end)
    if emp:
        query = query.filter(Employee.name.ilike(f"%{emp}%"))

    rows = query.order_by(Attendance.work_date.asc(), Employee.name.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    headers = ['Date', 'Code', 'Name', 'Dept', 'Check-In', 'Check-Out', 'Hours', 'Status', 'Note',
               'CheckIn IP', 'CheckIn UA', 'CheckOut IP', 'CheckOut UA']
    ws.append(headers)

    for r in rows:
        ws.append([
            r.work_date.isoformat(),
            r.employee.code,
            r.employee.name,
            r.employee.dept or '',
            to_local_iso(r.check_in),
            to_local_iso(r.check_out),
            r.duration_hours,
            r.status,
            r.note or '',
            r.check_in_ip or '',
            r.check_in_ua or '',
            r.check_out_ip or '',
            r.check_out_ua or '',
        ])

    # autosize columns (simple)
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        as_attachment=True,
        download_name="attendance_export.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.get("/leave/request")
@login_required
def leave_request_form():
    employees = Employee.query.filter_by(is_active=True).order_by(Employee.name.asc()).all()
    return render_template("leave_request.html", employees=employees)

@app.post("/leave/request")
@login_required
def leave_request_submit():
    emp_id = int(request.form.get("employee_id"))
    typ = request.form.get("type")
    start_date = datetime.strptime(request.form.get("start_date"), "%Y-%m-%d").date()
    end_date = datetime.strptime(request.form.get("end_date"), "%Y-%m-%d").date()
    reason = request.form.get("reason")

    # Initial status: staff -> pending_manager, manager -> pending_hrd, hrd/admin -> approved
    role = (current_user.role or "staff").lower()
    if role in ("admin", "hrd"):
        init_status = "approved"
    elif role in ("manager", "general_manager"):
        init_status = "pending_hrd"
    else:
        init_status = "pending_manager"

    lr = LeaveRequest(
        employee_id=emp_id,
        type=typ,
        start_date=start_date,
        end_date=end_date,
        reason=reason,
        status=init_status
    )
    db.session.add(lr)
    db.session.commit()

    # If created as approved (HRD/admin submitting), also mark attendance right away
    if init_status == "approved":
        lr.hrd_approved_by = current_user.id
        lr.hrd_approved_at = datetime.utcnow()

        d = lr.start_date
        while d <= lr.end_date:
            a = Attendance.query.filter_by(employee_id=lr.employee_id, work_date=d).first()
            if not a:
                a = Attendance(employee_id=lr.employee_id, work_date=d)
                db.session.add(a)
            a.status = "wfh" if lr.type == "wfh" else lr.type
            a.note = (a.note or "") + f" Marked by approval #{lr.id}."
            d += timedelta(days=1)

        db.session.commit()

    flash("Request submitted", "success")
    return redirect(url_for("leave_request_form"))

@app.get("/leave/approvals")
@login_required
def leave_approvals():
    role = (current_user.role or "staff").lower()

    # staff can still open the page, but it will only show their own requests
    if role == "staff":
        emp = _employee_for_user(current_user)
        if not emp:
            rows = []
        else:
            rows = LeaveRequest.query.filter_by(employee_id=emp.id).order_by(LeaveRequest.created_at.desc()).all()
        return render_template("approvals.html", rows=rows)

    # manager: only pending_manager for their department
    if role == "manager":
        me_emp = _employee_for_user(current_user)
        if not me_emp or not me_emp.dept:
            rows = []
        else:
            rows = (LeaveRequest.query
                    .join(Employee)
                    .filter(Employee.dept == me_emp.dept)
                    .filter(LeaveRequest.status == "pending_manager")
                    .order_by(LeaveRequest.created_at.desc())
                    .all())
        return render_template("approvals.html", rows=rows)

    # general manager: can see all pending_manager
    if role == "general_manager":
        rows = LeaveRequest.query.filter(LeaveRequest.status == "pending_manager").order_by(LeaveRequest.created_at.desc()).all()
        return render_template("approvals.html", rows=rows)

    # HRD: only pending_hrd
    if role == "hrd":
        rows = LeaveRequest.query.filter(LeaveRequest.status == "pending_hrd").order_by(LeaveRequest.created_at.desc()).all()
        return render_template("approvals.html", rows=rows)

    # admin: see all
    rows = LeaveRequest.query.order_by(LeaveRequest.created_at.desc()).all()
    return render_template("approvals.html", rows=rows)


@app.post("/leave/<int:rid>/approve")
@login_required
def leave_approve(rid):
    role = (current_user.role or "staff").lower()
    r = LeaveRequest.query.get_or_404(rid)

    # Manager stage
    if role in ("manager", "general_manager"):
        me_emp = _employee_for_user(current_user)
        if role == "manager":
            # manager only for same dept
            if not me_emp or not me_emp.dept or r.employee.dept != me_emp.dept:
                flash("Not allowed (different department).", "danger")
                return redirect(url_for("leave_approvals"))

        if r.status != "pending_manager":
            flash("Request is not waiting for manager approval.", "warning")
            return redirect(url_for("leave_approvals"))

        r.status = "pending_hrd"
        r.manager_approved_by = current_user.id
        r.manager_approved_at = datetime.utcnow()
        db.session.commit()
        flash("Approved (forwarded to HRD).", "success")
        return redirect(url_for("leave_approvals"))

    # HRD stage (final approve)
    if role in ("hrd", "admin"):
        if r.status not in ("pending_hrd", "pending_manager"):
            flash("Request is not waiting for HRD approval.", "warning")
            return redirect(url_for("leave_approvals"))

        # If admin approves directly from pending_manager, treat as HRD final approve
        r.status = "approved"
        r.hrd_approved_by = current_user.id
        r.hrd_approved_at = datetime.utcnow()
        db.session.commit()

        # Mark attendance for date range
        d = r.start_date
        while d <= r.end_date:
            a = Attendance.query.filter_by(employee_id=r.employee_id, work_date=d).first()
            if not a:
                a = Attendance(employee_id=r.employee_id, work_date=d)
                db.session.add(a)
            a.status = "wfh" if r.type == "wfh" else r.type
            a.note = (a.note or "") + f" Marked by approval #{r.id}."
            d += timedelta(days=1)

        db.session.commit()
        flash("Approved (final).", "success")
        return redirect(url_for("leave_approvals"))

    flash("Not allowed.", "danger")
    return redirect(url_for("dashboard"))


@app.post("/leave/<int:rid>/reject")
@login_required
def leave_reject(rid):
    role = (current_user.role or "staff").lower()
    r = LeaveRequest.query.get_or_404(rid)

    # Who can reject:
    # - manager/general_manager can reject pending_manager
    # - hrd/admin can reject pending_hrd (and admin can reject anything)
    allowed = False
    if role == "admin":
        allowed = True
    elif role in ("manager", "general_manager") and r.status == "pending_manager":
        allowed = True
        if role == "manager":
            me_emp = _employee_for_user(current_user)
            if not me_emp or not me_emp.dept or r.employee.dept != me_emp.dept:
                allowed = False
    elif role == "hrd" and r.status == "pending_hrd":
        allowed = True

    if not allowed:
        flash("Not allowed to reject this request.", "danger")
        return redirect(url_for("leave_approvals"))

    r.status = "rejected"
    r.rejected_by = current_user.id
    r.rejected_at = datetime.utcnow()
    db.session.commit()
    flash("Rejected.", "info")
    return redirect(url_for("leave_approvals"))


@app.post("/seed")
@login_required
def seed():
    if current_user.role != "admin":
        flash("Admins only", "warning")
        return redirect(url_for("dashboard"))
    if Employee.query.count() == 0:
        s = Shift.query.filter_by(name="Office").first()
        for i in range(1, 6):
            db.session.add(Employee(code=f"E{i:03d}", name=f"Employee {i}", email=f"user{i}@example.com", dept="OPS", shift_id=(s.id if s else None)))
        db.session.commit()
        flash("Seeded 5 employees", "success")
    else:
        flash("Employees already exist", "info")
    return redirect(url_for("employees"))

@app.route("/admin/offices", methods=["GET", "POST"])
@login_required
def offices_admin():
    if current_user.role != "admin":
        flash("Admins only", "warning")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        if "create" in request.form:
            db.session.add(Office(
                name=request.form["name"],
                lat=float(request.form["lat"]),
                lon=float(request.form["lon"]),
                radius_m=int(request.form["radius_m"]),
                is_active=True
            ))
            db.session.commit()
            flash("Office added", "success")

        elif "set_active" in request.form:
            Office.query.update({Office.is_active: False})
            Office.query.filter_by(id=int(request.form["office_id"])).update({Office.is_active: True})
            db.session.commit()
            flash("Active office updated", "success")

    offices = Office.query.order_by(Office.name.asc()).all()
    active = Office.query.filter_by(is_active=True).first()
    return render_template("office_admin.html", offices=offices, active=active)

@app.get("/admin/announcements")
@login_required
def announcements_admin():
    if current_user.role != "admin":
        flash("Admins only", "warning")
        return redirect(url_for("dashboard"))

    rows = Announcement.query.order_by(Announcement.created_at.desc()).all()
    return render_template("announcements.html", rows=rows)


@app.post("/admin/announcements/create")
@login_required
def announcement_create():
    if current_user.role != "admin":
        flash("Admins only", "warning")
        return redirect(url_for("dashboard"))

    title = (request.form.get("title") or "").strip()
    body = (request.form.get("body") or "").strip()
    level = (request.form.get("level") or "info").strip().lower()
    is_active = (request.form.get("is_active") == "1")

    # optional window
    start_at_s = (request.form.get("start_at") or "").strip()  # "YYYY-MM-DDTHH:MM"
    end_at_s = (request.form.get("end_at") or "").strip()

    def parse_dt_local(s):
        if not s:
            return None
        # input type=datetime-local -> "YYYY-MM-DDTHH:MM"
        try:
            return datetime.strptime(s, "%Y-%m-%dT%H:%M")
        except Exception:
            return None

    start_at = parse_dt_local(start_at_s)
    end_at = parse_dt_local(end_at_s)

    if not title or not body:
        flash("Title and body are required.", "danger")
        return redirect(url_for("announcements_admin"))

    if level not in ("info", "warning", "danger"):
        level = "info"

    a = Announcement(
        title=title,
        body=body,
        level=level,
        is_active=is_active,
        start_at=start_at,
        end_at=end_at,
    )
    db.session.add(a)
    db.session.commit()
    flash("Announcement added", "success")
    return redirect(url_for("announcements_admin"))


@app.post("/admin/announcements/<int:aid>/toggle")
@login_required
def announcement_toggle(aid):
    if current_user.role != "admin":
        flash("Admins only", "warning")
        return redirect(url_for("dashboard"))

    a = Announcement.query.get_or_404(aid)
    a.is_active = not bool(a.is_active)
    db.session.commit()
    flash("Announcement updated", "success")
    return redirect(url_for("announcements_admin"))


@app.post("/admin/announcements/<int:aid>/delete")
@login_required
def announcement_delete(aid):
    if current_user.role != "admin":
        flash("Admins only", "warning")
        return redirect(url_for("dashboard"))

    a = Announcement.query.get_or_404(aid)
    db.session.delete(a)
    db.session.commit()
    flash("Announcement deleted", "info")
    return redirect(url_for("announcements_admin"))

@app.get("/mobile")
def mobile_web():
    return render_template("mobile_app.html")


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(host="0.0.0.0", port=5000)
